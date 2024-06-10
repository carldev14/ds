from pymongo import MongoClient
from prompt_toolkit import prompt
from openpyxl.styles import Font, Alignment
from prompt_toolkit.completion import WordCompleter
from openpyxl import load_workbook
import os
import datetime
from dotenv import load_dotenv
load_dotenv()

# MongoDB connection
mongodb_url = os.getenv("MONGODB_URL")
client = MongoClient(mongodb_url)
mongodb_DB = os.getenv("DATABASE")
db = client[mongodb_DB]

companyCol = os.getenv("CompanyCol")
addressCol = os.getenv("AddressCol")
descriptionCol = os.getenv("DescriptionCol")

companies = db[companyCol]
addresses = db[addressCol]
descriptions = db[descriptionCol]

print('\nWelcome to Excel Automate. Use xlsx file for better experience.')
# Function to get the File PATH
def FilePath():
    while True:
        excel_path = prompt('\nEnter the PATH of excel file with extension: ')
        if not os.path.isfile(excel_path):
            print('\nFile not found')
        else: 
            open_workbook = load_workbook(excel_path)
            break
    return open_workbook, excel_path


def parse_date(input_str: str) -> tuple:
    # Remove any dots in the input string
    input_str = input_str.replace('.', '')
    
    # Extract month, day, and year components from the input string
    month = int(input_str[:2])
    day = int(input_str[2:4])
    year = int(input_str[4:])
    
    # Determine the correct century for the year component
    if year >= 0 and year <= 21:
        year += 2000
    else:
        year += 1900
    
    # Calculate the last two digits of the year
    year %= 100
    
    # Get the abbreviated month name
    month_abbr = datetime.date(year, month, day).strftime("%b")
    
    # Format the date string
    formatted_date = f"{day}-{month_abbr}-{year}"
    
    return formatted_date, input_str, month



def ChoiceSheet(month, workbook):
    month_names = {
        '01': 'January', '02': 'February', '03': 'March', '04': 'April',
        '05': 'May', '06': 'June', '07': 'July', '08': 'August',
        '09': 'September', '10': 'October', '11': 'November', '12': 'December'
    }
    sheet_name = month_names.get(month)
    if sheet_name:
        worksheet = workbook[sheet_name]
        return worksheet
    else:
        print('Not Found')

# Function to fetch suggestions from MongoDB
def fetch_company_suggestions():
    suggestions = companies.distinct("name")
    return suggestions

def fetch_address_suggestions():
    suggestions = addresses.distinct("name")
    return suggestions

def fetch_description_suggestions():
    suggestions = descriptions.distinct("name")
    return suggestions

def Autocomplete_suggestions(prompt_text, completer):
    print(f'\n{prompt_text}\n')
    # Prompt user for input with autocomplete
    while True:
        input_value = prompt(f"Enter the {prompt_text}> ", completer=completer)
        if input_value:
            return input_value

def Company_name():
    completer = WordCompleter(fetch_company_suggestions(), ignore_case=True)
    return Autocomplete_suggestions('Company Name', completer)

def Address_name():
    completer = WordCompleter(fetch_address_suggestions(), ignore_case=True)
    return Autocomplete_suggestions('Address', completer)

def Description_name():
    completer = WordCompleter(fetch_description_suggestions(), ignore_case=True)
    return Autocomplete_suggestions('Description', completer)

def TIN():
    while True:
        tin_number = prompt('\nEnter the TIN number: ')
        if tin_number == "":
            print('\nFill the blank\n')
        else:
            TIN = tin_number
            return TIN
        
def Save_to_excel(workbook, worksheet, excel_path, formatted_date, company_named, address_named, tin_number_input, description_named, TotalAmountPaid, InputVat, CPS):

    # Get the next available row
    next_row = worksheet.max_row + 1

    # Set values for the new row
    font = Font(name="Arial", size=10)
    alignment = Alignment(horizontal='center', vertical='center')

    row_data = [
        formatted_date,
        company_named,  # Corrected variable name
        address_named,  # Corrected variable name
        tin_number_input,  # Corrected variable name
        description_named,  # Corrected variable name
        TotalAmountPaid,
        InputVat,
        CPS
    ]

    worksheet.append(row_data)

    # Update font and alignment for each cell in the new row
    for cell in worksheet[next_row]:
        cell.font = font
        cell.alignment = alignment

    # Save the workbook
    workbook.save(excel_path)  # Corrected variable name

    print("\033[92mData has been successfully saved.\033[00m")


def Again(workbook, worksheet, excel_path, formatted_date, company_named, address_named, tin_number_input, description_named, TotalAmountPaid, InputVat, CPS):
    while True:
        ask = input('\nDo you create more? (y/n): ')
        if ask.lower() == 'y':
            Input_details(prompt('\nEnter the date (052324 = May,24,2024): '))
        else:
            print('GoodBye!')
            exit()   
             
def Input_details(input_str):
    try:
        formatted_date, input_str, month = parse_date(input_str)
    except ValueError:
        print("\nInvalid date format. Please enter the date in the format MMDDYY.")
        return
    company_named = Company_name()
    address_named = Address_name()
    description_named = Description_name()
    tin_number_input = TIN()
    TotalAmountPaid = float(prompt('\nEnter the Total Amount Paid: '))
    InputVat = float(prompt('\nEnter the Input Vat: '))
    CPS = TotalAmountPaid - InputVat 
    worksheet = ChoiceSheet(str(month).zfill(2), workbook)
    
    
    #Confirm Details
    
    print('\nConfirm Details\n')
    print(f'Date: {formatted_date}')
    print(f'Company: {company_named}')
    print(f'Address: {address_named}')
    print(f'TIN: {tin_number_input}')
    print(f'Description: {description_named}')
    print(f':Total Amount Paid: {TotalAmountPaid}')
    print(f'Input Vat: {InputVat}')
    print(f'Cost of Product and Services: {CPS} ')
    
    while True:
        y_or_n = input('\nDoes the details are correct? (y/n): ')
        if y_or_n.lower() == 'y':
            Save_to_excel(workbook, worksheet, excel_path, formatted_date, company_named, address_named, tin_number_input, description_named, TotalAmountPaid, InputVat, CPS)
            Again(workbook, worksheet, excel_path, formatted_date, company_named, address_named, tin_number_input, description_named, TotalAmountPaid, InputVat, CPS)
            break
        else:
            Input_details(prompt('\nEnter the date (052324 = May,24,2024): '))
            Again(workbook, worksheet, excel_path, formatted_date, company_named, address_named, tin_number_input, description_named, TotalAmountPaid, InputVat, CPS)
            break


if __name__ == "__main__":
    workbook, excel_path = FilePath()
    Input_details(prompt('Enter the date (052324 = May,24,2024): '))
